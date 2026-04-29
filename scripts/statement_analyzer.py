#!/usr/bin/env python3
"""
scripts/statement_analyzer.py
==============================
Parse an ICICI savings-account PDF statement and produce an Excel report
with categorised income / expense analysis.

Usage:
    python scripts/statement_analyzer.py <pdf_path> [-o output.xlsx]
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from collections import defaultdict
from datetime import datetime

import pdfplumber
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, BarChart3D
from openpyxl.chart.label import DataLabelList

# Make Unicode prints (₹, ✅, →) work on Windows cp1252 consoles.
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass


# ── Categorisation rules ────────────────────────────────────────────────────
# Each rule = (regex, category, subcategory).  First match wins.
CATEGORY_RULES: list[tuple[str, str, str]] = [
    # ── Income ─────────────────────────────────────────────────────────────
    (r"INTEL\s*TECHNOLOGY.*REIM",                "Income", "Salary Reimbursement"),
    (r"INTEL\s*TECHNOLOGY",                      "Income", "Salary"),
    (r"NEFT-.*INTEL",                            "Income", "Salary"),
    (r"\bSALARY\b|\bSAL\b",                      "Income", "Salary"),
    (r"INTEREST\s*PAID|SAVINGS\s*INTEREST|CR\s*INT|SB\s*INT", "Income", "Bank Interest"),
    (r"REFUND|REVERSAL",                         "Income", "Refund"),
    (r"CASHBACK|REWARD",                         "Income", "Cashback / Reward"),
    (r"DIVIDEND",                                "Income", "Investment Income"),
    # ── Investments / EMI ──────────────────────────────────────────────────
    (r"ZERODHA|GROWW|UPSTOX|KITE|COIN\b|MF\s*UTIL|ICCL|BSE\s*LTD|NSE\s*CLEARING|INDIAN\s*CLEARING",
                                                 "Investments", "Stocks / Broker"),
    (r"SIP|MUTUAL\s*FUND|MF\s*INVEST|KUVERA|PAYTM\s*MONEY",
                                                 "Investments", "Mutual Fund / SIP"),
    (r"PPF|EPF|NPS\b|KISAN\s*VIKAS|SUKANYA",     "Investments", "Govt. / Retirement"),
    (r"FD\b|FIXED\s*DEPOSIT|RD\b|RECURRING\s*DEPOSIT", "Investments", "FD / RD"),
    (r"GOLDBOND|SGB|GOLD\s*ETF|DIGITAL\s*GOLD|SAFEGOLD|MMTC",
                                                 "Investments", "Gold"),
    (r"LIC\s*OF\s*INDIA|LIFE\s*INSURANCE|HDFC\s*LIFE|ICICI\s*PRU|TATA\s*AIA",
                                                 "Insurance", "Life Insurance"),
    (r"HEALTH\s*INSURANCE|STAR\s*HEALTH|NIVA|MAX\s*BUPA|RELIGARE|MEDICLAIM",
                                                 "Insurance", "Health Insurance"),
    (r"CAR\s*INSURANCE|VEHICLE\s*INSURANCE|MOTOR\s*INSURANCE|BAJAJ\s*ALLIANZ",
                                                 "Insurance", "Vehicle Insurance"),
    (r"\bEMI\b|LOAN\s*EMI|HOME\s*LOAN|HOUSING\s*LOAN", "Loans / EMI", "EMI"),
    (r"CREDIT\s*CARD|CCBILL|CARD\s*PAYMENT|AMAZONPAYCC|ONECARD|SLICEIT",
                                                 "Loans / EMI", "Credit Card Payment"),
    (r"CHEQ1@|YESBANK.*CHEQ",                    "Loans / EMI", "Credit Card Payment"),
    # ── Bills / Utilities ──────────────────────────────────────────────────
    (r"ELECTRIC|TNEB|BESCOM|MSEB|TANGEDCO|POWER\s*BILL", "Bills / Utilities", "Electricity"),
    (r"GAS|INDANE|HP\s*GAS|BHARAT\s*GAS",        "Bills / Utilities", "Gas / LPG"),
    (r"WATER|METROWATER|CMWSSB",                 "Bills / Utilities", "Water"),
    (r"BROADBAND|ACT\s*FIBER|JIO\s*FIBER|AIRTEL\s*XSTREAM|EXCITEL|HATHWAY",
                                                 "Bills / Utilities", "Broadband"),
    (r"AIRTEL|JIO\b|VODAFONE|VI\s*PREPAID|MOBILE\s*RECHARGE|RECHARGE",
                                                 "Bills / Utilities", "Mobile Recharge"),
    (r"DTH|TATASKY|TATA\s*PLAY|DISHTV|D2H",      "Bills / Utilities", "DTH / Cable"),
    (r"NETFLIX|HOTSTAR|PRIME\s*VIDEO|SONYLIV|ZEE5|YOUTUBE\s*PREMIUM|SPOTIFY|JIOSAAVN|GAANA",
                                                 "Subscriptions", "Streaming"),
    (r"APPLE\.COM|ICLOUD|GOOGLE\s*ONE|MICROSOFT\s*365|OFFICE\s*365|ADOBE",
                                                 "Subscriptions", "Software / Cloud"),
    (r"MAINTENANCE|APARTMENT|FLAT\s*MAINT|SOCIETY|MYGATE|NOBROKER\s*HOOD",
                                                 "Housing", "Maintenance / Society"),
    (r"\bRENT\b|HOUSE\s*RENT|RENTAL",            "Housing", "Rent"),
    # ── Food & groceries ───────────────────────────────────────────────────
    (r"ZOMATO",                                  "Food & Dining", "Zomato"),
    (r"SWIGGY\s*INSTAMART",                      "Groceries", "Swiggy Instamart"),
    (r"SWIGGY",                                  "Food & Dining", "Swiggy"),
    (r"EATSURE|FAASOS|BEHROUZ|FRESHMENU|BOX8|REBEL\s*FOOD",
                                                 "Food & Dining", "Cloud Kitchen"),
    (r"DOMINOS|PIZZAHUT|PIZZA\s*HUT|MCDONALD|KFC|BURGER\s*KING|SUBWAY|TACO\s*BELL",
                                                 "Food & Dining", "Fast Food (QSR)"),
    (r"BIGBASKET|BB\s*DAILY|BBNOW",              "Groceries", "BigBasket"),
    (r"BLINKIT|GROFERS",                         "Groceries", "Blinkit"),
    (r"ZEPTO",                                   "Groceries", "Zepto"),
    (r"DUNZO|JIOMART|INSTAMART",                 "Groceries", "Quick Commerce"),
    (r"DMART|D-MART|MORE\s*RETAIL|RELIANCE\s*FRESH|SPENCER|NILGIRIS|NATURE.*BASKET",
                                                 "Groceries", "Supermarket"),
    (r"BAKERY|HOTEL|RESTAURANT|DHABA|CAFE|COFFEE|BARISTA|STARBUCKS|CCD|CHAI\s*POINT|CHAAYOS",
                                                 "Food & Dining", "Restaurant / Cafe"),
    (r"POORI|MOMOS|DOSA|IDLI|BIRYANI|TIFFIN|MESS\b",
                                                 "Food & Dining", "Local Eatery"),
    # ── Travel / fuel ──────────────────────────────────────────────────────
    (r"OLA\b",                                   "Transport", "Ola"),
    (r"UBER",                                    "Transport", "Uber"),
    (r"RAPIDO",                                  "Transport", "Rapido"),
    (r"MERU|YULU|VOGO|BOUNCE",                   "Transport", "Bike / Other"),
    (r"IRCTC|RAILWAY",                           "Travel", "Train"),
    (r"MAKEMYTRIP|GOIBIBO|YATRA|CLEARTRIP|EASEMYTRIP|IXIGO",
                                                 "Travel", "Online Travel Agent"),
    (r"REDBUS|ABHIBUS|KSRTC|MSRTC|TNSTC|APSRTC", "Travel", "Bus"),
    (r"INDIGO|VISTARA|AIR\s*INDIA|SPICEJET|GOAIR|GOFIRST|AKASA|EMIRATES|FLIGHT",
                                                 "Travel", "Flights"),
    (r"OYO|AIRBNB|TREEBO|FAB\s*HOTEL|GOIBIBO\s*HOTEL", "Travel", "Hotel / Stay"),
    (r"FUEL|PETROL|HPCL|IOCL|BPCL|INDIANOIL|SHELL|ESSAR|RELIANCE\s*PETROL",
                                                 "Transport", "Fuel"),
    (r"FASTAG|TOLL|PAYTM\s*FASTAG|ICICI\s*FASTAG|NHAI",
                                                 "Transport", "Toll / Fastag"),
    (r"PARKING",                                 "Transport", "Parking"),
    # ── Shopping ───────────────────────────────────────────────────────────
    (r"AMAZON",                                  "Shopping", "Amazon"),
    (r"FLIPKART",                                "Shopping", "Flipkart"),
    (r"MYNTRA|AJIO|MEESHO|TATACLIQ|SNAPDEAL",    "Shopping", "Fashion Online"),
    (r"NYKAA",                                   "Shopping", "Beauty / Personal Care"),
    (r"CROMA|RELIANCE\s*DIGITAL|VIJAY\s*SALES",  "Shopping", "Electronics"),
    (r"DECATHLON|LIFESTYLE|SHOPPERS\s*STOP|MAX\s*FASHION|PANTALOONS|H\s*&\s*M|ZARA|UNIQLO|WESTSIDE",
                                                 "Shopping", "Apparel / Retail"),
    (r"IKEA|HOMECENTRE|PEPPERFRY|URBAN\s*LADDER", "Shopping", "Furniture / Home"),
    # ── Health / wellness ──────────────────────────────────────────────────
    (r"APOLLO|PHARM|MEDLIFE|1MG|TATA\s*1MG|NETMEDS|PHARMEASY|MEDPLUS|WELLNESS\s*FOREVER",
                                                 "Health", "Pharmacy"),
    (r"HOSPITAL|CLINIC|DOCTOR|DIAGNOSTIC|\bLAB\b|PRACTO|TATA\s*HEALTH",
                                                 "Health", "Hospital / Clinic"),
    (r"GYM|CULTFIT|CULT\.FIT|FITNESS|YOGA",      "Health", "Fitness / Gym"),
    (r"SALON|BARBER|HAIRCUT|URBANCLAP|URBANCOMPANY",
                                                 "Personal Care", "Salon / Grooming"),
    # ── Education / kids ───────────────────────────────────────────────────
    (r"SCHOOL|TUITION|FEES|VIDYA|EDUCATION|UDEMY|COURSERA|UPGRAD|BYJU|UNACADEMY|VEDANTU|TOPPR",
                                                 "Education", "Education"),
    (r"FIRSTCRY|HOPSCOTCH|MOTHERCARE",           "Family", "Kids"),
    # ── Donations / gifts ──────────────────────────────────────────────────
    (r"DONATION|TEMPLE|TIRUPATI|ISKCON|NGO|CRY\b|GIVEINDIA",
                                                 "Donations", "Donation / Religious"),
    (r"GIFT\b|GIFTING|CARDS\b",                  "Gifts", "Gift"),
    # ── Cash / transfers ───────────────────────────────────────────────────
    (r"ATM\s*CASH|CASH\s*WDL|NWD\s*ATM|CWDR|ATM\b",
                                                 "Cash", "ATM Withdrawal"),
    (r"NEFT\b|RTGS\b|IMPS\b",                    "Bank Transfer", "NEFT / IMPS / RTGS"),
    (r"UPI/.*@",                                 "UPI Transfer", "UPI Transfer"),
    (r"ACH/|ECS\b",                              "Auto-Debit", "ACH / ECS"),
    # ── Charges / taxes ────────────────────────────────────────────────────
    (r"GST\b|TDS\b|\bTAX\b|STAMP\s*DUTY|INCOME\s*TAX",
                                                 "Tax", "Tax / GST"),
    (r"CHARGE|FEE|PENALTY|LATE\s*PAYMENT|MIN\s*BAL|NON\s*MAINTENANCE",
                                                 "Bank Charges", "Bank Fee / Penalty"),
]

INCOME_CATEGORIES = {"Income"}

# Categories considered "essential" — not flagged as savings opportunities.
ESSENTIAL_CATEGORIES = {
    "Bills / Utilities", "Housing", "Health", "Education", "Insurance",
    "Loans / EMI", "Tax", "Investments", "Family",
}


def categorise(remark: str) -> tuple[str, str]:
    """Return (category, subcategory) for a transaction remark."""
    r = remark.upper()
    for pattern, cat, sub in CATEGORY_RULES:
        if re.search(pattern, r):
            return cat, sub
    return "Other", "Uncategorised"


# ── Counter-party identity tags ─────────────────────────────────────────────
# Optional JSON file maps a regex / substring → friendly label + relation.
# Example file: data/statement_tags.json
# {
#   "tags": [
#     {"match": "abhi.*@oksbi",          "name": "Abhishek (Brother)", "relation": "Family"},
#     {"match": "amma.upi@ybl",          "name": "Mother",             "relation": "Family"},
#     {"match": "INTEL\\s*TECHNOLOGY",   "name": "Intel Salary",       "relation": "Self"},
#     {"match": "ICIC000000001366478",   "name": "Union Bank Loan",    "relation": "Loan"}
#   ]
# }
DEFAULT_TAGS_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "data", "statement_tags.json",
)


def load_tags(path: str | None) -> list[dict]:
    p = path or DEFAULT_TAGS_PATH
    if not os.path.isfile(p):
        return []
    try:
        with open(p, "r", encoding="utf-8") as fh:
            data = json.load(fh)
        tags = data.get("tags", []) if isinstance(data, dict) else data
        return [t for t in tags if "match" in t]
    except Exception as exc:
        print(f"[warn] could not read tags file {p}: {exc}", file=sys.stderr)
        return []


def apply_tag(remark: str, tags: list[dict]) -> tuple[str, str]:
    """Return (counterparty, relation) — both '' if no tag matches."""
    if not tags:
        return "", ""
    r_up = remark.upper()
    for t in tags:
        pat = t.get("match", "")
        try:
            if re.search(pat, remark, re.I) or re.search(pat, r_up, re.I):
                return t.get("name", ""), t.get("relation", "")
        except re.error:
            if pat.upper() in r_up:
                return t.get("name", ""), t.get("relation", "")
    return "", ""


def extract_bank_counterparty(remark: str) -> str:
    """Best-effort extraction of the OTHER party for NEFT/IMPS/RTGS/ACH rows."""
    r = remark
    # NEFT: "NEFT-CITIN52025042955555305-INTEL TECHNOLOGY..."
    m = re.match(r"NEFT-([A-Z0-9]+)-([A-Z0-9 .,&/]+)", r)
    if m:
        return m.group(2).strip()[:60]
    # IMPS / RTGS: "IMPS/<ref>/<NAME>/..." or similar
    m = re.search(r"(?:IMPS|RTGS)[/\-:]([^/\-]+)[/\-]([^/]+)", r, re.I)
    if m:
        return m.group(2).strip()[:60]
    # ACH: "ACH/UNIONBANKOFINDIA/ICIC000000001366478/03199535"
    m = re.match(r"ACH/([^/]+)/", r)
    if m:
        return m.group(1).strip()[:60]
    return r[:60]


# ── Parser ──────────────────────────────────────────────────────────────────

# Matches a transaction row line:
#  <sno> <dd.mm.yyyy> [<cheque>] <amount> <balance>
# Cheque numbers are usually absent; when present they're 6+ digits w/o decimal.
_ROW_RE = re.compile(
    r"^\s*(\d+)\s+"                 # 1. S.No.
    r"(\d{2}\.\d{2}\.\d{4})\s+"     # 2. Date
    r"(?:(\d{6,})\s+)?"             # 3. (optional) cheque #
    r"([\d,]+\.\d{2})\s+"           # 4. Amount (withdrawal OR deposit)
    r"([\d,]+\.\d{2})\s*$"          # 5. Balance
)


def _to_float(s: str) -> float:
    return float(s.replace(",", ""))


def parse_pdf(path: str) -> list[dict]:
    """Parse the PDF and return a list of transaction dicts."""
    transactions: list[dict] = []
    prev_balance: float | None = None

    with pdfplumber.open(path) as pdf:
        # Concatenate every page's lines in order
        all_lines: list[str] = []
        for page in pdf.pages:
            txt = page.extract_text() or ""
            all_lines.extend(txt.splitlines())

    # Build the index of row-line positions, then attach surrounding
    # non-row lines as remarks.
    row_idxs: list[tuple[int, re.Match]] = []
    for i, ln in enumerate(all_lines):
        m = _ROW_RE.match(ln)
        if m:
            row_idxs.append((i, m))

    for k, (i, m) in enumerate(row_idxs):
        sno      = int(m.group(1))
        date_str = m.group(2)
        cheque   = m.group(3) or ""
        amount   = _to_float(m.group(4))
        balance  = _to_float(m.group(5))

        # Remarks: line immediately before this row (if it isn't another row
        # or a known header), plus every subsequent line up to the next row
        # line. ICICI's PDF puts the FIRST line of the remark above the row
        # (because the cell wraps) and the rest below.
        remark_parts: list[str] = []
        # one line above
        if i - 1 >= 0:
            prev = all_lines[i - 1].strip()
            if prev and not _ROW_RE.match(prev) and not prev.startswith(
                ("Statement of", "Transaction", "S No", "Page ", "Total", "Legends")
            ) and not prev.isdigit():
                remark_parts.append(prev)
        # lines below until next row
        end = row_idxs[k + 1][0] if k + 1 < len(row_idxs) else len(all_lines)
        for j in range(i + 1, end):
            ln = all_lines[j].strip()
            if not ln or ln.isdigit():
                continue
            if ln.startswith(("Page ", "Statement of", "Legends")):
                continue
            remark_parts.append(ln)

        remark = " ".join(remark_parts).strip()

        # Classify debit vs credit using balance delta
        if prev_balance is None:
            # First row: use heuristic — if amount ≈ balance, treat as opening
            # credit; otherwise default credit.
            kind = "credit"
        else:
            delta = round(balance - prev_balance, 2)
            if abs(delta + amount) < 0.05:
                kind = "debit"
            elif abs(delta - amount) < 0.05:
                kind = "credit"
            else:
                # Fallback: sign of delta
                kind = "credit" if delta >= 0 else "debit"
        prev_balance = balance

        try:
            d = datetime.strptime(date_str, "%d.%m.%Y").date()
        except ValueError:
            continue

        cat, sub = categorise(remark)
        transactions.append({
            "sno":      sno,
            "date":     d,
            "cheque":   cheque,
            "remark":   remark,
            "withdrawal": amount if kind == "debit"  else 0.0,
            "deposit":    amount if kind == "credit" else 0.0,
            "balance":  balance,
            "category": cat,
            "subcategory": sub,
        })

    return transactions


# ── Excel report ────────────────────────────────────────────────────────────

def _merchant(remark: str) -> str:
    """Pull a friendly merchant token from a UPI / NEFT remark."""
    m = re.match(r"UPI/([^/]+)/([^/]+)", remark)
    if m:
        handle = m.group(1).strip()
        note   = m.group(2).strip()
        return f"UPI: {handle} ({note[:30]})" if note and note.upper() != "UPI" else f"UPI: {handle}"
    m = re.match(r"NEFT-([A-Z0-9]+).*?-([A-Z0-9 &.,]+)", remark)
    if m:
        return f"NEFT: {m.group(2).strip()[:40]}"
    m = re.match(r"ACH/([A-Z0-9 &.]+)", remark)
    if m:
        return f"ACH: {m.group(1).strip()[:40]}"
    return remark[:60]


def _detect_recurring(df: "pd.DataFrame") -> "pd.DataFrame":
    """
    A merchant is 'recurring' if it appears in ≥3 distinct months with a
    similar amount each time (median-deviation ≤ 25 %). Useful to surface
    forgotten subscriptions.
    """
    out_rows = []
    for merchant, sub in df.groupby("merchant"):
        if len(sub) < 3:
            continue
        months = sub["month"].nunique()
        if months < 3:
            continue
        amts = sub["withdrawal"].abs()
        med  = amts.median()
        if med <= 0:
            continue
        dev  = (amts - med).abs().median() / med
        if dev > 0.25:
            continue
        out_rows.append({
            "merchant":   merchant,
            "category":   sub["category"].mode().iat[0],
            "months":     months,
            "txns":       len(sub),
            "avg_amount": round(float(amts.mean()), 2),
            "total":      round(float(amts.sum()),  2),
            "last_seen":  sub["date"].max(),
        })
    rec = pd.DataFrame(out_rows).sort_values("total", ascending=False) if out_rows else pd.DataFrame(
        columns=["merchant", "category", "months", "txns", "avg_amount", "total", "last_seen"]
    )
    return rec


def _build_savings_plan(df: "pd.DataFrame", n_months: int) -> "pd.DataFrame":
    """
    Walk through known 'leaky' categories and estimate a realistic monthly
    saving if usage were trimmed.  Each row carries the rationale.
    """
    plans: list[dict] = []
    exp = df[df["type"] == "Expense"].copy()

    def add(area, current_pm, target_pm, action):
        save_pm = max(0.0, current_pm - target_pm)
        plans.append({
            "area":              area,
            "current_per_month": round(current_pm, 0),
            "target_per_month":  round(target_pm,  0),
            "potential_saving_per_month": round(save_pm, 0),
            "potential_saving_per_year":  round(save_pm * 12, 0),
            "action":            action,
        })

    # 1. Food delivery (Zomato + Swiggy + cloud kitchens)
    food_apps = exp[exp["subcategory"].isin(
        ["Zomato", "Swiggy", "Cloud Kitchen", "Fast Food (QSR)"])]
    if not food_apps.empty:
        per_m = food_apps["withdrawal"].sum() / n_months
        # Target: cap at ₹4,000/month (3-4 orders) or 50 % of current, whichever lower
        tgt = min(per_m * 0.5, 4000)
        add("Food delivery (Zomato/Swiggy/QSR)", per_m, tgt,
            "Cap to 4 orders/month; cook 2 weekday dinners; share orders.")

    # 2. Quick-commerce groceries (Blinkit / Zepto / Instamart)
    qc = exp[exp["subcategory"].isin(["Blinkit", "Zepto", "Quick Commerce", "Swiggy Instamart"])]
    if not qc.empty:
        per_m = qc["withdrawal"].sum() / n_months
        tgt   = per_m * 0.4
        add("Quick-commerce impulse buys", per_m, tgt,
            "Plan a weekly grocery list at DMart/BigBasket; use Blinkit only for true emergencies.")

    # 3. Cafes / restaurants
    rest = exp[exp["subcategory"].isin(["Restaurant / Cafe", "Local Eatery"])]
    if not rest.empty:
        per_m = rest["withdrawal"].sum() / n_months
        tgt   = per_m * 0.6
        add("Cafes / dining out", per_m, tgt,
            "Limit cafe visits to once a week; carry coffee/snacks from home.")

    # 4. Cabs (Ola/Uber)
    cab = exp[exp["subcategory"].isin(["Ola", "Uber", "Rapido", "Bike / Other"])]
    if not cab.empty:
        per_m = cab["withdrawal"].sum() / n_months
        tgt   = per_m * 0.5
        add("Ride-hailing (Ola/Uber/Rapido)", per_m, tgt,
            "Switch to metro/bus for routine trips; pool with colleagues.")

    # 5. Online shopping
    shop = exp[exp["category"] == "Shopping"]
    if not shop.empty:
        per_m = shop["withdrawal"].sum() / n_months
        tgt   = per_m * 0.6
        add("Online shopping", per_m, tgt,
            "Adopt a 48-hour 'cart-cooling' rule; uninstall shopping apps; remove saved cards.")

    # 6. Streaming + software subs
    sub = exp[exp["category"] == "Subscriptions"]
    if not sub.empty:
        per_m = sub["withdrawal"].sum() / n_months
        tgt   = per_m * 0.5
        add("Streaming & software subscriptions", per_m, tgt,
            "Audit recurring list (see 'Recurring Charges' sheet); keep 1-2 OTT, share family plans.")

    # 7. ATM cash
    atm = exp[exp["subcategory"] == "ATM Withdrawal"]
    if not atm.empty:
        per_m = atm["withdrawal"].sum() / n_months
        tgt   = per_m * 0.5
        add("ATM cash leakage", per_m, tgt,
            "Cash spent has no trail. Withdraw less; track every ₹500 of cash use in a notes app.")

    # 8. Bank charges / fees
    fee = exp[exp["category"] == "Bank Charges"]
    if not fee.empty:
        per_m = fee["withdrawal"].sum() / n_months
        add("Bank charges / penalties", per_m, 0,
            "Avoid late payments, maintain min balance, use free ATMs — these should be ZERO.")

    # 9. Bottom 50 % small UPI transfers (< ₹300, person-to-person)
    petty = exp[(exp["subcategory"] == "UPI Transfer") & (exp["withdrawal"] < 300)]
    if len(petty) > 30:
        per_m = petty["withdrawal"].sum() / n_months
        tgt   = per_m * 0.5
        add("Petty UPI spends (<₹300)", per_m, tgt,
            f"Found {len(petty)} small payments; combine errands and budget ₹X/week of pocket cash.")

    return pd.DataFrame(plans)


def _build_insights(df: "pd.DataFrame", n_months: int,
                    total_income: float, total_expense: float) -> list[str]:
    """Generate plain-English bullet observations."""
    notes: list[str] = []
    exp = df[df["type"] == "Expense"]
    inc = df[df["type"] == "Income"]

    # 1. Savings rate vs ideal 30 %
    rate = (total_income - total_expense) / max(1, total_income) * 100
    if rate < 20:
        notes.append(f"⚠ Savings rate is {rate:.1f}% — well below the 20–30 % healthy band. "
                     f"Aim to lift it to ≥20 % over the next 3 months.")
    elif rate < 30:
        notes.append(f"Savings rate is {rate:.1f}% — okay, but the 30 % target is within reach "
                     f"with the cuts in the 'Savings Plan' sheet.")
    else:
        notes.append(f"✅ Savings rate is {rate:.1f}% — strong. Focus next on tax-efficient investing.")

    # 2. Top 3 expense categories
    topcats = (exp.groupby("category")["withdrawal"].sum()
                  .sort_values(ascending=False).head(3))
    for cat, amt in topcats.items():
        share = amt / max(1, total_expense) * 100
        notes.append(f"• Largest spend area: {cat} — ₹{amt:,.0f} ({share:.1f}% of expense, "
                     f"₹{amt/n_months:,.0f}/month).")

    # 3. Discretionary vs essential split
    ess = exp[exp["category"].isin(ESSENTIAL_CATEGORIES)]["withdrawal"].sum()
    disc = total_expense - ess
    notes.append(f"• Essential vs discretionary: ₹{ess:,.0f} essential ({ess/total_expense*100:.0f}%)  vs  "
                 f"₹{disc:,.0f} discretionary ({disc/total_expense*100:.0f}%). "
                 f"Discretionary is where most savings live.")

    # 4. Late-night / weekend spend spotlight
    df_e = exp.copy()
    df_e["dow"] = pd.to_datetime(df_e["date"]).dt.dayofweek
    weekend = df_e[df_e["dow"] >= 5]["withdrawal"].sum()
    if weekend > 0:
        wk_pm = weekend / n_months
        notes.append(f"• Weekend spend averages ₹{wk_pm:,.0f}/month "
                     f"({weekend/total_expense*100:.0f}% of total). "
                     f"Weekends are the biggest impulse-window — pre-plan Sat/Sun budgets.")

    # 5. Food-delivery spotlight
    fd = exp[exp["subcategory"].isin(["Zomato", "Swiggy", "Cloud Kitchen", "Fast Food (QSR)"])]
    if not fd.empty:
        fd_pm = fd["withdrawal"].sum() / n_months
        fd_orders = len(fd) / n_months
        notes.append(f"• Food-delivery avg: ₹{fd_pm:,.0f}/month across ~{fd_orders:.0f} orders/month. "
                     f"Avg order value ₹{fd['withdrawal'].mean():,.0f}.")

    # 6. Quick-commerce spotlight
    qc = exp[exp["subcategory"].isin(["Blinkit", "Zepto", "Quick Commerce", "Swiggy Instamart"])]
    if not qc.empty:
        qc_pm = qc["withdrawal"].sum() / n_months
        notes.append(f"• Quick-commerce (Blinkit/Zepto/Instamart): ₹{qc_pm:,.0f}/month — "
                     f"these typically carry 15-30 % markup vs supermarkets.")

    # 7. ATM cash spotlight
    atm = exp[exp["subcategory"] == "ATM Withdrawal"]
    if not atm.empty:
        atm_pm = atm["withdrawal"].sum() / n_months
        notes.append(f"• ATM cash: ₹{atm_pm:,.0f}/month — cash transactions have no audit trail; "
                     f"reduce by 50 % and switch to UPI for traceability.")

    # 8. Bank charges
    fee = exp[exp["category"] == "Bank Charges"]
    if not fee.empty:
        notes.append(f"⚠ Paid ₹{fee['withdrawal'].sum():,.0f} in bank fees / penalties — "
                     f"these are 100 % avoidable.")

    # 9. Investment vs expense
    invs = exp[exp["category"] == "Investments"]["withdrawal"].sum()
    inv_share = invs / max(1, total_income) * 100
    if inv_share < 15:
        notes.append(f"• Active investments are only {inv_share:.1f}% of income. "
                     f"Target ≥20 % through SIPs (mutual funds, NPS, PPF).")
    else:
        notes.append(f"✅ Investments are {inv_share:.1f}% of income — solid.")

    # 10. High-value outliers
    big = exp[exp["withdrawal"] >= 25000].sort_values("withdrawal", ascending=False).head(3)
    if not big.empty:
        items = ", ".join(f"₹{r['withdrawal']:,.0f} ({r['subcategory']})"
                          for _, r in big.iterrows())
        notes.append(f"• Top 3 single expenses: {items}. Review whether these were planned.")

    return notes


def write_excel(transactions: list[dict], out_path: str,
                tags: list[dict] | None = None) -> None:
    df = pd.DataFrame(transactions)
    if df.empty:
        raise SystemExit("No transactions parsed — check the PDF format.")

    tags = tags or []
    df["month"]  = pd.to_datetime(df["date"]).dt.to_period("M").astype(str)
    df["amount"] = df["deposit"] - df["withdrawal"]
    df["type"]   = "Expense"
    df.loc[df["deposit"] > 0, "type"] = "Income"
    df["merchant"] = df["remark"].map(_merchant)
    df["weekday"]  = pd.to_datetime(df["date"]).dt.day_name()

    # Apply counter-party tags
    tag_results = df["remark"].map(lambda r: apply_tag(r, tags))
    df["counterparty"] = [t[0] for t in tag_results]
    df["relation"]     = [t[1] for t in tag_results]
    # Best-effort counterparty for bank-transfer rows when no tag matched
    is_bt = df["category"].isin(["Bank Transfer", "Loans / EMI"]) | \
            df["remark"].str.contains(r"^(?:NEFT|IMPS|RTGS|ACH)", case=False, regex=True)
    df.loc[is_bt & (df["counterparty"] == ""), "counterparty"] = \
        df.loc[is_bt & (df["counterparty"] == ""), "remark"].map(extract_bank_counterparty)

    # ── Headline aggregates ──────────────────────────────────────────────
    total_income  = df.loc[df["type"] == "Income",  "deposit"].sum()
    total_expense = df.loc[df["type"] == "Expense", "withdrawal"].sum()
    net_savings   = total_income - total_expense
    n_txn         = len(df)
    period_start  = df["date"].min()
    period_end    = df["date"].max()
    n_months      = max(1, df["month"].nunique())

    exp_df = df[df["type"] == "Expense"].copy()
    inc_df = df[df["type"] == "Income"].copy()

    essential_total      = exp_df[exp_df["category"].isin(ESSENTIAL_CATEGORIES)]["withdrawal"].sum()
    discretionary_total  = total_expense - essential_total

    summary_rows = [
        ("Statement period",      f"{period_start}  →  {period_end}"),
        ("Months covered",        n_months),
        ("Transactions",          n_txn),
        ("Total income (₹)",      round(float(total_income),  2)),
        ("Total expense (₹)",     round(float(total_expense), 2)),
        ("Net savings (₹)",       round(float(net_savings),   2)),
        ("Avg monthly income",    round(float(total_income)  / n_months, 2)),
        ("Avg monthly expense",   round(float(total_expense) / n_months, 2)),
        ("Avg monthly savings",   round(float(net_savings)   / n_months, 2)),
        ("Savings rate (%)",      round(float(net_savings)   / float(total_income or 1) * 100, 1)),
        ("Essential expense (₹)",     round(float(essential_total), 2)),
        ("Discretionary expense (₹)", round(float(discretionary_total), 2)),
        ("Discretionary share (%)",
            round(float(discretionary_total) / float(total_expense or 1) * 100, 1)),
    ]
    summary_df = pd.DataFrame(summary_rows, columns=["Metric", "Value"])

    # ── Insights & savings plan ──────────────────────────────────────────
    insights = _build_insights(df, n_months, float(total_income), float(total_expense))
    insights_df = pd.DataFrame({"#": list(range(1, len(insights) + 1)), "Observation": insights})

    savings_plan_df = _build_savings_plan(df, n_months)
    if not savings_plan_df.empty:
        total_save_pm = savings_plan_df["potential_saving_per_month"].sum()
        total_save_yr = savings_plan_df["potential_saving_per_year"].sum()
        total_row = pd.DataFrame([{
            "area": "TOTAL POTENTIAL",
            "current_per_month": "",
            "target_per_month":  "",
            "potential_saving_per_month": total_save_pm,
            "potential_saving_per_year":  total_save_yr,
            "action": "Implement top 3 areas first; track monthly.",
        }])
        savings_plan_df = pd.concat([savings_plan_df, total_row], ignore_index=True)

    # ── Category and Subcategory rollups ─────────────────────────────────
    by_cat = (df.groupby(["type", "category"])
                .agg(amount=("amount", lambda s: float(abs(s).sum())),
                     count=("amount", "size"),
                     avg=("amount", lambda s: float(abs(s).mean())))
                .reset_index()
                .sort_values(["type", "amount"], ascending=[True, False]))
    expense_total_local = by_cat.loc[by_cat["type"] == "Expense", "amount"].sum() or 1
    income_total_local  = by_cat.loc[by_cat["type"] == "Income",  "amount"].sum() or 1
    by_cat["share_%"] = by_cat.apply(
        lambda r: round(r["amount"] /
                        (expense_total_local if r["type"] == "Expense" else income_total_local) * 100, 1),
        axis=1,
    )
    by_cat["per_month"] = (by_cat["amount"] / n_months).round(0)

    by_sub = (exp_df.groupby(["category", "subcategory"])
                    .agg(amount=("withdrawal", "sum"),
                         count=("withdrawal", "size"),
                         avg=("withdrawal", "mean"),
                         max=("withdrawal", "max"))
                    .reset_index()
                    .sort_values("amount", ascending=False))
    by_sub["share_%"]   = (by_sub["amount"] / expense_total_local * 100).round(1)
    by_sub["per_month"] = (by_sub["amount"] / n_months).round(0)

    # ── Monthly cashflow & category × month ──────────────────────────────
    monthly = (df.groupby(["month", "type"])["amount"]
                 .apply(lambda s: float(abs(s).sum()))
                 .unstack(fill_value=0.0)
                 .reset_index())
    if "Income"  not in monthly.columns: monthly["Income"]  = 0.0
    if "Expense" not in monthly.columns: monthly["Expense"] = 0.0
    monthly["Net"] = monthly["Income"] - monthly["Expense"]
    monthly["Savings_%"] = (monthly["Net"] / monthly["Income"].replace(0, pd.NA) * 100).round(1)
    monthly = monthly[["month", "Income", "Expense", "Net", "Savings_%"]]

    cat_pivot = (exp_df.pivot_table(index="category", columns="month",
                                    values="withdrawal", aggfunc="sum",
                                    fill_value=0.0).round(2))
    cat_pivot["Total"] = cat_pivot.sum(axis=1).round(2)
    cat_pivot = cat_pivot.sort_values("Total", ascending=False).reset_index()

    sub_pivot = (exp_df.pivot_table(index=["category", "subcategory"], columns="month",
                                    values="withdrawal", aggfunc="sum",
                                    fill_value=0.0).round(2))
    sub_pivot["Total"] = sub_pivot.sum(axis=1).round(2)
    sub_pivot = sub_pivot.sort_values("Total", ascending=False).reset_index()

    # ── Day-of-week pattern ──────────────────────────────────────────────
    dow_df = (exp_df.groupby("weekday")["withdrawal"]
                    .agg(["sum", "count", "mean"])
                    .reset_index()
                    .rename(columns={"sum": "total", "count": "txns", "mean": "avg"}))
    weekday_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    dow_df["weekday"] = pd.Categorical(dow_df["weekday"], categories=weekday_order, ordered=True)
    dow_df = dow_df.sort_values("weekday").reset_index(drop=True)
    dow_df[["total", "avg"]] = dow_df[["total", "avg"]].round(2)

    # ── Top merchants & top transactions ────────────────────────────────
    top_merch = (exp_df.groupby(["merchant", "category"])
                       .agg(total=("withdrawal", "sum"),
                            txns=("withdrawal", "size"),
                            avg=("withdrawal", "mean"),
                            first=("date", "min"),
                            last=("date", "max"))
                       .reset_index()
                       .sort_values("total", ascending=False)
                       .head(40))
    top_merch[["total", "avg"]] = top_merch[["total", "avg"]].round(2)

    top_txn = (df.assign(abs_amt=df["amount"].abs())
                 .sort_values("abs_amt", ascending=False)
                 .head(30)
                 [["date", "type", "category", "subcategory", "remark",
                   "withdrawal", "deposit", "balance"]])

    # ── Recurring / subscription detector ────────────────────────────────
    recurring_df = _detect_recurring(exp_df)

    # ── Small / petty UPI dump (the 'invisible bleed') ───────────────────
    petty = (exp_df[(exp_df["subcategory"] == "UPI Transfer") &
                    (exp_df["withdrawal"] < 300)]
                .groupby("merchant")
                .agg(total=("withdrawal", "sum"),
                     txns=("withdrawal", "size"),
                     avg=("withdrawal", "mean"))
                .reset_index()
                .sort_values("total", ascending=False).head(30))
    if not petty.empty:
        petty[["total", "avg"]] = petty[["total", "avg"]].round(2)

    # ── Bank-transfer drill-down (NEFT / IMPS / RTGS / ACH) ──────────────
    bt_mask = df["category"].isin(["Bank Transfer", "Loans / EMI"]) | \
              df["remark"].str.contains(r"^(?:NEFT|IMPS|RTGS|ACH)", case=False, regex=True)
    bt_df = df[bt_mask].copy()
    bt_df["counterparty_label"] = bt_df.apply(
        lambda r: r["counterparty"] if r["counterparty"] else extract_bank_counterparty(r["remark"]),
        axis=1,
    )
    bt_drill = (bt_df.groupby(["counterparty_label", "type", "relation"], dropna=False)
                     .agg(total=("amount", lambda s: float(abs(s).sum())),
                          txns=("amount", "size"),
                          avg=("amount",  lambda s: float(abs(s).mean())),
                          first=("date",  "min"),
                          last=("date",   "max"))
                     .reset_index()
                     .sort_values("total", ascending=False))
    bt_drill[["total", "avg"]] = bt_drill[["total", "avg"]].round(2)
    bt_drill["per_month"] = (bt_drill["total"] / n_months).round(0)

    # ── Spend grouped by tagged relation (Family / Self / Loan / etc.) ──
    by_relation_df = pd.DataFrame()
    if (df["relation"] != "").any():
        rel = df[df["relation"] != ""].copy()
        by_relation_df = (rel.groupby(["relation", "type"])
                             .agg(total=("amount", lambda s: float(abs(s).sum())),
                                  txns=("amount", "size"))
                             .reset_index()
                             .sort_values(["type", "total"], ascending=[True, False]))
        by_relation_df["total"] = by_relation_df["total"].round(2)
        by_relation_df["per_month"] = (by_relation_df["total"] / n_months).round(0)

    # ── Write workbook ───────────────────────────────────────────────────
    out_path = os.path.abspath(out_path)
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as xl:
        summary_df.to_excel(xl,      sheet_name="Summary",            index=False)
        insights_df.to_excel(xl,     sheet_name="Insights",           index=False)
        if not savings_plan_df.empty:
            savings_plan_df.to_excel(xl, sheet_name="Savings Plan",   index=False)
        by_cat.to_excel(xl,          sheet_name="By Category",        index=False)
        by_sub.to_excel(xl,          sheet_name="By Subcategory",     index=False)
        if not by_relation_df.empty:
            by_relation_df.to_excel(xl, sheet_name="By Relation",     index=False)
        monthly.to_excel(xl,         sheet_name="Monthly Cashflow",   index=False)
        cat_pivot.to_excel(xl,       sheet_name="Category x Month",   index=False)
        sub_pivot.to_excel(xl,       sheet_name="Subcategory x Month", index=False)
        dow_df.to_excel(xl,          sheet_name="Day of Week",        index=False)
        if not bt_drill.empty:
            bt_drill.to_excel(xl,    sheet_name="Bank Transfer Drill", index=False)
        if not recurring_df.empty:
            recurring_df.to_excel(xl, sheet_name="Recurring Charges",  index=False)
        if not petty.empty:
            petty.to_excel(xl,        sheet_name="Petty UPI",          index=False)
        top_merch.to_excel(xl,       sheet_name="Top Merchants",      index=False)
        top_txn.to_excel(xl,         sheet_name="Top Transactions",   index=False)
        df_export = df[["date", "weekday", "category", "subcategory", "type",
                        "counterparty", "relation",
                        "merchant", "remark",
                        "withdrawal", "deposit", "balance", "cheque"]]
        df_export.to_excel(xl,       sheet_name="All Transactions",   index=False)

        # ── Pivot-style charts ───────────────────────────────────────────
        wb = xl.book

        # 1. Pie chart of Expense By Category (top 12)
        ws = wb["By Category"]
        n_exp = int((by_cat["type"] == "Expense").sum())
        if n_exp >= 2:
            n_show = min(12, n_exp)
            pie = PieChart()
            labels = Reference(ws, min_col=2, min_row=2, max_row=1 + n_show)        # category col
            data   = Reference(ws, min_col=3, min_row=1, max_row=1 + n_show)        # amount col w/ header
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.title = "Expense share by category (top categories)"
            pie.height = 12
            pie.width  = 18
            pie.dataLabels = DataLabelList(showPercent=True)
            ws.add_chart(pie, "I2")

            bar = BarChart()
            bar.type = "bar"
            bar.style = 11
            bar.title = "Expense (₹) by category"
            bar.y_axis.title = "Category"
            bar.x_axis.title = "₹ total"
            bar.add_data(data, titles_from_data=True)
            bar.set_categories(labels)
            bar.height = 12
            bar.width  = 18
            ws.add_chart(bar, "I26")

        # 2. By Subcategory: top-15 horizontal bar
        ws = wb["By Subcategory"]
        n_sub = min(15, len(by_sub))
        if n_sub >= 2:
            labels = Reference(ws, min_col=2, min_row=2, max_row=1 + n_sub)         # subcategory
            data   = Reference(ws, min_col=3, min_row=1, max_row=1 + n_sub)         # amount header+data
            bar = BarChart()
            bar.type = "bar"
            bar.style = 12
            bar.title = "Top 15 subcategories by spend"
            bar.add_data(data, titles_from_data=True)
            bar.set_categories(labels)
            bar.height = 14
            bar.width  = 22
            ws.add_chart(bar, "J2")

        # 3. Monthly Cashflow: combined bar (Income+Expense) + line (Net)
        ws = wb["Monthly Cashflow"]
        n_m = len(monthly)
        if n_m >= 2:
            cats = Reference(ws, min_col=1, min_row=2, max_row=1 + n_m)             # month
            inc_exp = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=1 + n_m)  # Income, Expense
            net_col = Reference(ws, min_col=4, min_row=1, max_row=1 + n_m)             # Net
            bar = BarChart()
            bar.type = "col"
            bar.style = 10
            bar.title = "Monthly Income vs Expense (with Net line)"
            bar.y_axis.title = "₹"
            bar.x_axis.title = "Month"
            bar.add_data(inc_exp, titles_from_data=True)
            bar.set_categories(cats)
            bar.height = 12
            bar.width  = 24
            try:
                line = LineChart()
                line.add_data(net_col, titles_from_data=True)
                line.y_axis.axId = 200
                line.y_axis.crosses = "max"
                bar += line
            except Exception:
                pass
            ws.add_chart(bar, "G2")

            # Savings_% line chart
            sav_pct = Reference(ws, min_col=5, min_row=1, max_row=1 + n_m)
            ln = LineChart()
            ln.title = "Monthly savings rate (%)"
            ln.add_data(sav_pct, titles_from_data=True)
            ln.set_categories(cats)
            ln.y_axis.title = "%"
            ln.x_axis.title = "Month"
            ln.height = 10
            ln.width  = 24
            ws.add_chart(ln, "G24")

        # 4. Day-of-week bar
        ws = wb["Day of Week"]
        if len(dow_df) >= 2:
            cats = Reference(ws, min_col=1, min_row=2, max_row=1 + len(dow_df))
            data = Reference(ws, min_col=2, min_row=1, max_row=1 + len(dow_df))
            bar = BarChart()
            bar.type = "col"
            bar.style = 13
            bar.title = "Spend by day of week"
            bar.add_data(data, titles_from_data=True)
            bar.set_categories(cats)
            bar.height = 10
            bar.width  = 20
            ws.add_chart(bar, "F2")

        # 5. Savings Plan bar chart
        if "Savings Plan" in wb.sheetnames and not savings_plan_df.empty:
            ws = wb["Savings Plan"]
            n_sp = len(savings_plan_df) - 1   # exclude TOTAL row
            if n_sp >= 1:
                cats = Reference(ws, min_col=1, min_row=2, max_row=1 + n_sp)
                data = Reference(ws, min_col=4, min_row=1, max_row=1 + n_sp)  # potential_saving_per_month
                bar = BarChart()
                bar.type = "bar"
                bar.style = 14
                bar.title = "Potential ₹ savings per month by area"
                bar.add_data(data, titles_from_data=True)
                bar.set_categories(cats)
                bar.height = 12
                bar.width  = 22
                ws.add_chart(bar, "H2")

        # 6. Category × Month: stacked column chart (top 8 categories)
        ws = wb["Category x Month"]
        if len(cat_pivot) >= 2 and len(monthly) >= 2:
            n_cat_show = min(8, len(cat_pivot))
            n_cols = len(cat_pivot.columns)   # category, m1..mN, Total
            cats = Reference(ws, min_col=1, min_row=2, max_row=1 + n_cat_show)   # category names
            # Use month columns (skip first = category, last = Total)
            data = Reference(ws, min_col=2, max_col=n_cols - 1,
                             min_row=1, max_row=1 + n_cat_show)
            bar = BarChart()
            bar.type = "col"
            bar.grouping = "stacked"
            bar.overlap = 100
            bar.style = 10
            bar.title = "Top categories: monthly stack"
            bar.add_data(data, titles_from_data=True)
            bar.set_categories(cats)
            bar.height = 14
            bar.width  = 26
            ws.add_chart(bar, f"{get_column_letter(n_cols + 2)}2")

        # 7. Bank Transfer Drill bar chart (top 15)
        if "Bank Transfer Drill" in wb.sheetnames and not bt_drill.empty:
            ws = wb["Bank Transfer Drill"]
            n_bt = min(15, len(bt_drill))
            cats = Reference(ws, min_col=1, min_row=2, max_row=1 + n_bt)
            data = Reference(ws, min_col=4, min_row=1, max_row=1 + n_bt)   # total
            bar = BarChart()
            bar.type = "bar"
            bar.style = 12
            bar.title = "Top 15 bank-transfer counter-parties (₹)"
            bar.add_data(data, titles_from_data=True)
            bar.set_categories(cats)
            bar.height = 14
            bar.width  = 22
            ws.add_chart(bar, "K2")

        # 8. Top Merchants chart
        ws = wb["Top Merchants"]
        n_tm = min(15, len(top_merch))
        if n_tm >= 2:
            cats = Reference(ws, min_col=1, min_row=2, max_row=1 + n_tm)
            data = Reference(ws, min_col=3, min_row=1, max_row=1 + n_tm)   # total
            bar = BarChart()
            bar.type = "bar"
            bar.style = 13
            bar.title = "Top 15 merchants by spend"
            bar.add_data(data, titles_from_data=True)
            bar.set_categories(cats)
            bar.height = 14
            bar.width  = 22
            ws.add_chart(bar, "I2")

        # 9. By Relation pie chart
        if "By Relation" in wb.sheetnames and not by_relation_df.empty:
            ws = wb["By Relation"]
            n_rel = len(by_relation_df)
            cats = Reference(ws, min_col=1, min_row=2, max_row=1 + n_rel)
            data = Reference(ws, min_col=3, min_row=1, max_row=1 + n_rel)
            pie = PieChart()
            pie.title = "Money flow by relation"
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(cats)
            pie.dataLabels = DataLabelList(showPercent=True)
            pie.height = 12
            pie.width  = 18
            ws.add_chart(pie, "G2")

        # Cosmetic: header style + autofilter + column widths
        header_fill = PatternFill("solid", fgColor="FFD700")
        header_font = Font(bold=True)
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for c in ws[1]:
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(horizontal="center", vertical="center")
            for col_idx, col in enumerate(ws.columns, start=1):
                max_len = 12
                for cell in col:
                    val = cell.value
                    if val is None:
                        continue
                    max_len = max(max_len, min(70, len(str(val)) + 2))
                ws.column_dimensions[get_column_letter(col_idx)].width = max_len

    # ── Console preview ──────────────────────────────────────────────────
    print(f"\n✅ Report written to: {out_path}")
    print(f"   Income:  ₹{total_income:>14,.2f}")
    print(f"   Expense: ₹{total_expense:>14,.2f}")
    print(f"   Net:     ₹{net_savings:>14,.2f}")
    print(f"   Period:  {period_start} → {period_end}  ({n_months} months, {n_txn} txns)")
    print(f"\n── KEY INSIGHTS ─────────────────────────────────────────────")
    for ln in insights:
        print("  " + ln)
    if not savings_plan_df.empty:
        print(f"\n── POTENTIAL MONTHLY SAVINGS ────────────────────────────────")
        for _, r in savings_plan_df.iterrows():
            sv = r["potential_saving_per_month"]
            if isinstance(sv, (int, float)) and sv:
                print(f"  • {r['area']:<40s} ₹{sv:>8,.0f}/mo")


# ── CLI ─────────────────────────────────────────────────────────────────────

def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Analyze ICICI bank PDF statement → Excel report.")
    p.add_argument("pdf", help="Path to the statement PDF")
    p.add_argument("-o", "--output", default=None,
                   help="Output xlsx path (default: alongside the PDF)")
    p.add_argument("--tags", default=None,
                   help=f"JSON file mapping counter-parties (default: {DEFAULT_TAGS_PATH})")
    return p


def main() -> None:
    args = _build_parser().parse_args()
    if not os.path.isfile(args.pdf):
        sys.exit(f"PDF not found: {args.pdf}")
    out = args.output or os.path.join(
        os.path.dirname(os.path.abspath(args.pdf)),
        os.path.splitext(os.path.basename(args.pdf))[0] + "_analysis.xlsx",
    )
    tags = load_tags(args.tags)
    if tags:
        print(f"Loaded {len(tags)} counter-party tags.")
    print(f"Parsing: {args.pdf}")
    txns = parse_pdf(args.pdf)
    print(f"Parsed {len(txns)} transactions.")
    write_excel(txns, out, tags=tags)


if __name__ == "__main__":
    main()
